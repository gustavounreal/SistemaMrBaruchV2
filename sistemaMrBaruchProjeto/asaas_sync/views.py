"""
Views para visualização de dados sincronizados do Asaas
"""
from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.db.models import Q, Sum, Count, Case, When, DecimalField
from django.utils import timezone
from django.conf import settings
from .models import AsaasClienteSyncronizado, AsaasCobrancaSyncronizada, AsaasSyncronizacaoLog
from .services import AsaasSyncService
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import logging
import subprocess
import threading
import sys

logger = logging.getLogger(__name__)


@login_required
def dashboard_asaas_sync(request):
    """Dashboard principal com dados do Asaas"""
    
    # Estatísticas gerais
    total_clientes = AsaasClienteSyncronizado.objects.count()
    total_cobrancas = AsaasCobrancaSyncronizada.objects.count()
    
    # Valores
    total_recebido = AsaasCobrancaSyncronizada.objects.filter(
        status__in=['RECEIVED', 'CONFIRMED']
    ).aggregate(total=Sum('valor'))['total'] or 0
    
    total_pendente = AsaasCobrancaSyncronizada.objects.filter(
        status='PENDING'
    ).aggregate(total=Sum('valor'))['total'] or 0
    
    total_vencido = AsaasCobrancaSyncronizada.objects.filter(
        status='OVERDUE'
    ).aggregate(total=Sum('valor'))['total'] or 0
    
    # Últimas sincronizações
    ultimas_syncs = AsaasSyncronizacaoLog.objects.all()[:5]
    ultima_sync = ultimas_syncs.first()
    
    # Cobranças recentes
    cobrancas_recentes = AsaasCobrancaSyncronizada.objects.select_related('cliente').order_by('-sincronizado_em')[:10]
    
    # Cobranças por status
    cobrancas_por_status = AsaasCobrancaSyncronizada.objects.values('status').annotate(
        total=Count('id'),
        valor_total=Sum('valor')
    ).order_by('-total')
    
    context = {
        'total_clientes': total_clientes,
        'total_cobrancas': total_cobrancas,
        'total_recebido': total_recebido,
        'total_pendente': total_pendente,
        'total_vencido': total_vencido,
        'ultimas_syncs': ultimas_syncs,
        'ultima_sync': ultima_sync,
        'cobrancas_recentes': cobrancas_recentes,
        'cobrancas_por_status': cobrancas_por_status,
    }
    
    return render(request, 'asaas_sync/dashboard.html', context)


@login_required
def lista_clientes(request):
    """Lista todos os clientes sincronizados com paginação"""
    from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
    
    # Filtros
    busca = request.GET.get('busca', '')
    inadimplente = request.GET.get('inadimplente', '')
    servico_concluido = request.GET.get('servico_concluido', '')
    page = request.GET.get('page', 1)
    
    clientes = AsaasClienteSyncronizado.objects.prefetch_related('cobrancas').all()
    
    # Filtro por nome, CPF ou CNPJ
    if busca:
        clientes = clientes.filter(
            Q(nome__icontains=busca) |
            Q(cpf_cnpj__icontains=busca)
        )
    
    # Filtro por inadimplência
    if inadimplente:
        if inadimplente == 'sim':
            # Clientes com cobranças vencidas
            clientes_ids = []
            for cliente in clientes:
                if cliente.esta_inadimplente():
                    clientes_ids.append(cliente.id)
            clientes = clientes.filter(id__in=clientes_ids)
        elif inadimplente == 'nao':
            # Clientes sem cobranças vencidas
            clientes_ids = []
            for cliente in clientes:
                if not cliente.esta_inadimplente():
                    clientes_ids.append(cliente.id)
            clientes = clientes.filter(id__in=clientes_ids)
    
    # Filtro por serviço concluído
    if servico_concluido:
        if servico_concluido == 'sim':
            clientes = clientes.filter(servico_concluido=True)
        elif servico_concluido == 'nao':
            clientes = clientes.filter(servico_concluido=False)
    
    # Ordenação
    clientes = clientes.order_by('nome')
    
    # Contagem total antes da paginação
    total_clientes = clientes.count()
    
    # Paginação - 50 clientes por página
    paginator = Paginator(clientes, 50)
    
    try:
        clientes_paginados = paginator.page(page)
    except PageNotAnInteger:
        clientes_paginados = paginator.page(1)
    except EmptyPage:
        clientes_paginados = paginator.page(paginator.num_pages)
    
    # Preparar dados enriquecidos
    clientes_dados = []
    for cliente in clientes_paginados:
        clientes_dados.append({
            'cliente': cliente,
            'valor_total': cliente.get_valor_total_servico(),
            'esta_inadimplente': cliente.esta_inadimplente(),
            'periodo_inadimplencia': cliente.get_periodo_inadimplencia(),
            'valor_inadimplente': cliente.get_valor_inadimplente(),
            'quantidade_boletos_pendentes': cliente.get_quantidade_boletos_pendentes(),
            'servicos_contratados': cliente.get_servicos_contratados(),
        })
    
    context = {
        'clientes_dados': clientes_dados,
        'clientes_paginados': clientes_paginados,
        'total_clientes': total_clientes,
        'busca': busca,
        'inadimplente': inadimplente,
        'servico_concluido': servico_concluido,
    }
    
    return render(request, 'asaas_sync/lista_clientes.html', context)


@login_required
def detalhes_cliente(request, cliente_id):
    """Detalhes de um cliente e suas cobranças"""
    
    cliente = get_object_or_404(AsaasClienteSyncronizado, id=cliente_id)
    
    # Cobranças do cliente
    cobrancas = cliente.cobrancas.all().order_by('-data_vencimento')
    
    # Estatísticas do cliente
    stats = {
        'total_cobrancas': cobrancas.count(),
        'total_recebido': cobrancas.filter(status__in=['RECEIVED', 'CONFIRMED']).aggregate(Sum('valor'))['valor__sum'] or 0,
        'total_pendente': cobrancas.filter(status='PENDING').aggregate(Sum('valor'))['valor__sum'] or 0,
        'total_vencido': cobrancas.filter(status='OVERDUE').aggregate(Sum('valor'))['valor__sum'] or 0,
    }
    
    # Separar cobranças por status
    cobrancas_pagas = cobrancas.filter(status__in=['RECEIVED', 'CONFIRMED'])
    cobrancas_pendentes = cobrancas.filter(status='PENDING')
    cobrancas_vencidas = cobrancas.filter(status='OVERDUE')
    cobrancas_outras = cobrancas.exclude(status__in=['RECEIVED', 'CONFIRMED', 'PENDING', 'OVERDUE'])
    
    context = {
        'cliente': cliente,
        'cobrancas': cobrancas,
        'stats': stats,
        'cobrancas_pagas': cobrancas_pagas,
        'cobrancas_pendentes': cobrancas_pendentes,
        'cobrancas_vencidas': cobrancas_vencidas,
        'cobrancas_outras': cobrancas_outras,
    }
    
    return render(request, 'asaas_sync/detalhes_cliente.html', context)


@login_required
def lista_cobrancas(request):
    """Lista todas as cobranças com filtros e paginação"""
    from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
    
    # Filtros
    status = request.GET.get('status', '')
    tipo = request.GET.get('tipo', '')
    cliente_busca = request.GET.get('cliente', '')
    data_inicio = request.GET.get('data_inicio', '')
    data_fim = request.GET.get('data_fim', '')
    page = request.GET.get('page', 1)
    
    cobrancas = AsaasCobrancaSyncronizada.objects.select_related('cliente').all()
    
    if status:
        cobrancas = cobrancas.filter(status=status)
    
    if tipo:
        cobrancas = cobrancas.filter(tipo_cobranca=tipo)
    
    if cliente_busca:
        cobrancas = cobrancas.filter(
            Q(cliente__nome__icontains=cliente_busca) |
            Q(cliente__cpf_cnpj__icontains=cliente_busca)
        )
    
    if data_inicio:
        cobrancas = cobrancas.filter(data_vencimento__gte=data_inicio)
    
    if data_fim:
        cobrancas = cobrancas.filter(data_vencimento__lte=data_fim)
    
    # Ordenação
    cobrancas = cobrancas.order_by('-data_vencimento')
    
    # Totais antes da paginação
    total_cobrancas = cobrancas.count()
    valor_total = cobrancas.aggregate(Sum('valor'))['valor__sum'] or 0
    
    # Paginação - 100 cobranças por página
    paginator = Paginator(cobrancas, 100)
    
    try:
        cobrancas_paginadas = paginator.page(page)
    except PageNotAnInteger:
        cobrancas_paginadas = paginator.page(1)
    except EmptyPage:
        cobrancas_paginadas = paginator.page(paginator.num_pages)
    
    # Totais
    totais = {
        'quantidade': total_cobrancas,
        'valor_total': valor_total
    }
    
    # Status filtrados para o dropdown
    status_filtrados = [
        ('PENDING', 'Pendente'),
        ('RECEIVED', 'Recebida'),
        ('CONFIRMED', 'Confirmada'),
        ('OVERDUE', 'Vencida'),
    ]
    
    context = {
        'cobrancas': cobrancas_paginadas,
        'cobrancas_paginadas': cobrancas_paginadas,
        'total_cobrancas': total_cobrancas,
        'status': status,
        'tipo': tipo,
        'cliente_busca': cliente_busca,
        'data_inicio': data_inicio,
        'data_fim': data_fim,
        'totais': totais,
        'status_choices': status_filtrados,
        'tipo_choices': AsaasCobrancaSyncronizada.TIPO_COBRANCA_CHOICES,
    }
    
    return render(request, 'asaas_sync/lista_cobrancas.html', context)


@login_required
def sincronizar_agora(request):
    """Inicia sincronização manual - TODOS os clientes e cobranças"""
    
    if request.method == 'POST':
        try:
            logger.info(f"Sincronização COMPLETA iniciada por {request.user.username}")
            
            sync_service = AsaasSyncService()
            log = sync_service.sincronizar_tudo(usuario=request.user, limite_clientes=None)  # None = TODOS
            
            return JsonResponse({
                'success': True,
                'message': 'Sincronização concluída com sucesso!',
                'log': {
                    'id': log.id,
                    'status': log.status,
                    'total_clientes': log.total_clientes,
                    'clientes_novos': log.clientes_novos,
                    'total_cobrancas': log.total_cobrancas,
                    'cobrancas_novas': log.cobrancas_novas,
                    'duracao': log.duracao_segundos,
                    'mensagem': log.mensagem,
                }
            })
            
        except Exception as e:
            logger.error(f"Erro na sincronização: {str(e)}", exc_info=True)
            return JsonResponse({
                'success': False,
                'message': f'Erro na sincronização: {str(e)}'
            }, status=500)
    
    return JsonResponse({
        'success': False,
        'message': 'Método não permitido'
    }, status=405)


@login_required
def sincronizar_boletos_faltantes(request):
    """
    Sincroniza APENAS os boletos faltantes dos clientes já cadastrados.
    Mostra progresso em tempo real.
    """
    
    if request.method == 'POST':
        try:
            from asaas_sync.models import AsaasSyncronizacaoLog
            from django.utils import timezone
            
            logger.info(f"Sincronização de BOLETOS FALTANTES iniciada por {request.user.username}")
            
            # Criar log inicial
            log = AsaasSyncronizacaoLog.objects.create(
                tipo_sincronizacao='BOLETOS_FALTANTES',
                status='EM_ANDAMENTO',
                usuario=request.user.username,
                mensagem='Sincronização iniciada...'
            )
            
            # Executar sincronização em thread separada
            import threading
            
            def executar_sync():
                try:
                    inicio = timezone.now()
                    sync_service = AsaasSyncService()
                    stats = sync_service.sincronizar_boletos_faltantes()
                    fim = timezone.now()
                    
                    # Atualizar log com resultados
                    log.data_fim = fim
                    log.status = 'SUCESSO'
                    log.total_clientes = stats.get('clientes_processados', 0)
                    log.cobrancas_novas = stats.get('cobrancas_novas', 0)
                    log.cobrancas_atualizadas = stats.get('cobrancas_atualizadas', 0)
                    log.total_cobrancas = stats.get('cobrancas_novas', 0) + stats.get('cobrancas_atualizadas', 0)
                    log.mensagem = f"Sincronização concluída com sucesso. {stats.get('clientes_processados', 0)} clientes processados."
                    log.calcular_duracao()
                    log.save()
                    
                    logger.info(f"✅ Sincronização de boletos concluída: {stats}")
                    
                except Exception as e:
                    log.data_fim = timezone.now()
                    log.status = 'ERRO'
                    log.erros = str(e)
                    log.mensagem = f'Erro durante sincronização: {str(e)}'
                    log.calcular_duracao()
                    log.save()
                    logger.error(f"Erro na sincronização de boletos: {str(e)}", exc_info=True)
            
            thread = threading.Thread(target=executar_sync)
            thread.daemon = True
            thread.start()
            
            return JsonResponse({
                'success': True,
                'message': 'Sincronização de boletos faltantes iniciada. Acompanhe o progresso na página.',
                'log_id': log.id
            })
            
        except Exception as e:
            logger.error(f"Erro ao iniciar sincronização de boletos: {str(e)}", exc_info=True)
            return JsonResponse({
                'success': False,
                'message': f'Erro ao iniciar sincronização: {str(e)}'
            }, status=500)
    
    return JsonResponse({
        'success': False,
        'message': 'Método não permitido'
    }, status=405)


@login_required
def acompanhar_sincronizacao(request, log_id):
    """Página de acompanhamento em tempo real da sincronização"""
    log = get_object_or_404(AsaasSyncronizacaoLog, id=log_id)
    
    context = {
        'log': log,
        'titulo': 'Acompanhamento de Sincronização',
    }
    
    return render(request, 'asaas_sync/acompanhar_sincronizacao.html', context)


@login_required
def status_sincronizacao(request, log_id):
    """API para obter status em tempo real da sincronização"""
    try:
        log = get_object_or_404(AsaasSyncronizacaoLog, id=log_id)
        
        return JsonResponse({
            'success': True,
            'status': log.status,
            'status_display': log.get_status_display(),
            'tipo': log.tipo_sincronizacao,
            'tipo_display': log.get_tipo_sincronizacao_display(),
            'mensagem': log.mensagem or '',
            'data_inicio': log.data_inicio.strftime('%d/%m/%Y %H:%M:%S') if log.data_inicio else '',
            'data_fim': log.data_fim.strftime('%d/%m/%Y %H:%M:%S') if log.data_fim else '',
            'duracao_segundos': log.duracao_segundos,
            'total_clientes': log.total_clientes,
            'clientes_novos': log.clientes_novos,
            'clientes_atualizados': log.clientes_atualizados,
            'total_cobrancas': log.total_cobrancas,
            'cobrancas_novas': log.cobrancas_novas,
            'cobrancas_atualizadas': log.cobrancas_atualizadas,
            'erros': log.erros or '',
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


@login_required
def historico_sincronizacoes(request):
    """Página com histórico de todas as sincronizações"""
    tipo_filtro = request.GET.get('tipo', '')
    
    logs = AsaasSyncronizacaoLog.objects.all().order_by('-data_inicio')
    
    if tipo_filtro:
        logs = logs.filter(tipo_sincronizacao=tipo_filtro)
    
    context = {
        'logs': logs,
        'tipo_filtro': tipo_filtro,
        'titulo': 'Histórico de Sincronizações',
    }
    
    return render(request, 'asaas_sync/historico_sincronizacoes.html', context)


@login_required
def relatorio_completo(request):
    """Relatório completo com todas as informações"""
    
    clientes = AsaasClienteSyncronizado.objects.all()
    
    dados = []
    
    for cliente in clientes:
        cobrancas = cliente.cobrancas.all()
        
        dados.append({
            'cliente': cliente,
            'total_cobrancas': cobrancas.count(),
            'cobrancas_pagas': cobrancas.filter(status__in=['RECEIVED', 'CONFIRMED']).count(),
            'cobrancas_pendentes': cobrancas.filter(status='PENDING').count(),
            'cobrancas_vencidas': cobrancas.filter(status='OVERDUE').count(),
            'valor_recebido': cobrancas.filter(status__in=['RECEIVED', 'CONFIRMED']).aggregate(Sum('valor'))['valor__sum'] or 0,
            'valor_pendente': cobrancas.filter(status='PENDING').aggregate(Sum('valor'))['valor__sum'] or 0,
            'valor_vencido': cobrancas.filter(status='OVERDUE').aggregate(Sum('valor'))['valor__sum'] or 0,
            'cobrancas': cobrancas.order_by('-data_vencimento')
        })
    
    context = {
        'dados': dados,
        'data_geracao': timezone.now(),
    }
    
    return render(request, 'asaas_sync/relatorio_completo.html', context)


@login_required
def atualizar_cliente(request, cliente_id):
    """Atualiza dados do cliente (consultor, status do serviço e serviços contratados)"""
    
    if request.method == 'POST':
        try:
            cliente = get_object_or_404(AsaasClienteSyncronizado, id=cliente_id)
            
            # Atualizar consultor
            if 'consultor_responsavel' in request.POST:
                consultor = request.POST.get('consultor_responsavel', '').strip()
                cliente.consultor_responsavel = consultor if consultor else None
            
            # Atualizar status do serviço
            if 'servico_concluido' in request.POST:
                servico = request.POST.get('servico_concluido', 'false')
                cliente.servico_concluido = servico.lower() == 'true'
            
            # Atualizar serviços contratados
            if 'servico_limpa_nome' in request.POST:
                cliente.servico_limpa_nome = request.POST.get('servico_limpa_nome') == 'true'
            if 'servico_retirada_travas' in request.POST:
                cliente.servico_retirada_travas = request.POST.get('servico_retirada_travas') == 'true'
            if 'servico_restauracao_score' in request.POST:
                cliente.servico_restauracao_score = request.POST.get('servico_restauracao_score') == 'true'
            
            cliente.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Cliente atualizado com sucesso!',
                'data': {
                    'consultor_responsavel': cliente.consultor_responsavel or '',
                    'servico_concluido': cliente.servico_concluido,
                    'servico_limpa_nome': cliente.servico_limpa_nome,
                    'servico_retirada_travas': cliente.servico_retirada_travas,
                    'servico_restauracao_score': cliente.servico_restauracao_score,
                    'servicos_display': cliente.get_servicos_contratados_display(),
                }
            })
            
        except Exception as e:
            logger.error(f"Erro ao atualizar cliente: {str(e)}", exc_info=True)
            return JsonResponse({
                'success': False,
                'message': f'Erro ao atualizar: {str(e)}'
            }, status=500)
    
    return JsonResponse({
        'success': False,
        'message': 'Método não permitido'
    }, status=405)


@login_required
def sincronizar_alternativo(request):
    """Sincroniza dados de uma conta Asaas alternativa usando processamento em background"""
    
    if request.method == 'POST':
        try:
            # Buscar token alternativo das configurações
            token_alternativo = getattr(settings, 'ASAAS_ALTERNATIVO_TOKEN', None)
            
            if not token_alternativo:
                return JsonResponse({
                    'success': False,
                    'message': '❌ Token alternativo não configurado.\n\nConfigure ASAAS_ALTERNATIVO_TOKEN nas variáveis de ambiente.'
                }, status=500)
            
            logger.info(f"Sincronização alternativa iniciada por {request.user.username}")
            
            # Iniciar sincronização em background usando comando Django
            def executar_sync():
                try:
                    # Caminho para o manage.py
                    base_dir = settings.BASE_DIR
                    manage_py = base_dir / 'manage.py'
                    
                    # Executar comando em background
                    python_executable = sys.executable
                    
                    comando = [
                        python_executable,
                        str(manage_py),
                        'sincronizar_asaas_alternativo'
                    ]
                    
                    logger.info(f"Executando comando: {' '.join(comando)}")
                    
                    # Executar sem bloquear
                    subprocess.Popen(
                        comando,
                        stdout=subprocess.PIPE,
                        stderr=subprocess.PIPE,
                        cwd=str(base_dir)
                    )
                    
                except Exception as e:
                    logger.error(f"Erro ao iniciar sincronização em background: {str(e)}", exc_info=True)
            
            # Iniciar thread para não bloquear a resposta HTTP
            thread = threading.Thread(target=executar_sync)
            thread.daemon = True
            thread.start()
            
            return JsonResponse({
                'success': True,
                'message': '🔄 Sincronização iniciada em background!\n\nO processo pode levar alguns minutos. Recarregue a página em instantes para ver os novos dados.'
            })
            
        except Exception as e:
            logger.error(f"Erro ao iniciar sincronização alternativa: {str(e)}", exc_info=True)
            return JsonResponse({
                'success': False,
                'message': f'❌ Erro ao iniciar sincronização: {str(e)}'
            }, status=500)
    
    return JsonResponse({
        'success': False,
        'message': 'Método não permitido'
    }, status=405)


@login_required
def validar_sincronizacao(request):
    """
    Valida a sincronização comparando dados do ASAAS com o banco local.
    Mostra discrepâncias entre o que existe no ASAAS e o que foi baixado.
    """
    
    # Verificar qual token usar
    usar_alternativo = request.GET.get('alternativo', 'false') == 'true'
    
    try:
        sync_service = AsaasSyncService()
        
        # Se usar alternativo, substituir token
        if usar_alternativo:
            token_alternativo = getattr(settings, 'ASAAS_ALTERNATIVO_TOKEN', None)
            if token_alternativo:
                sync_service.api_token = token_alternativo
                sync_service.headers['access_token'] = token_alternativo
                logger.info("Usando token ASAAS ALTERNATIVO para validação")
            else:
                return render(request, 'asaas_sync/validacao.html', {
                    'erro': 'Token alternativo não configurado'
                })
        
        # 1. VALIDAR CLIENTES - BUSCAR TODOS
        logger.info("Buscando TODOS os clientes do ASAAS...")
        
        todos_clientes_asaas = []
        offset = 0
        limit = 100
        
        while True:
            response_clientes = sync_service._fazer_requisicao('GET', 'customers', params={'limit': limit, 'offset': offset})
            
            if not response_clientes:
                # Se o token principal falhou, tentar o alternativo automaticamente
                if not usar_alternativo and offset == 0:
                    token_alternativo = getattr(settings, 'ASAAS_ALTERNATIVO_TOKEN', None)
                    if token_alternativo:
                        logger.warning("Token principal falhou, tentando token alternativo...")
                        sync_service.api_token = token_alternativo
                        sync_service.headers['access_token'] = token_alternativo
                        response_clientes = sync_service._fazer_requisicao('GET', 'customers', params={'limit': limit, 'offset': offset})
                        if response_clientes:
                            usar_alternativo = True
                            logger.info("✅ Token alternativo funcionou!")
                
                # Se ainda falhou, mostrar erro
                if not response_clientes:
                    return render(request, 'asaas_sync/validacao.html', {
                        'erro': 'Erro ao buscar clientes do ASAAS. Ambos os tokens falharam. Verifique as credenciais.',
                        'usar_alternativo': usar_alternativo
                    })
                else:
                    break
            
            clientes_data = response_clientes.get('data', [])
            todos_clientes_asaas.extend(clientes_data)
            
            total_count = response_clientes.get('totalCount', 0)
            has_more = response_clientes.get('hasMore', False)
            
            logger.info(f"Buscados {len(todos_clientes_asaas)}/{total_count} clientes...")
            
            if not has_more:
                break
            
            offset += limit
            
            # Limite de segurança
            if offset >= 10000:
                logger.warning("Limite de 10000 clientes atingido")
                break
        
        total_clientes_asaas = len(todos_clientes_asaas)
        logger.info(f"✅ Total de clientes no ASAAS: {total_clientes_asaas}")
        
        # Contar clientes no banco local
        total_clientes_local = AsaasClienteSyncronizado.objects.count()
        
        # Verificar quais clientes do ASAAS NÃO estão no banco
        clientes_faltando = []
        
        for cliente_data in todos_clientes_asaas:
            asaas_id = cliente_data.get('id')
            if not AsaasClienteSyncronizado.objects.filter(asaas_customer_id=asaas_id).exists():
                clientes_faltando.append({
                    'asaas_id': asaas_id,
                    'nome': cliente_data.get('name', 'N/A'),
                    'cpf_cnpj': cliente_data.get('cpfCnpj', 'N/A'),
                    'email': cliente_data.get('email', 'N/A'),
                })
        
        logger.info(f"Clientes faltando: {len(clientes_faltando)}")
        
        # 2. VALIDAR COBRANÇAS - TODOS OS CLIENTES
        logger.info("Validando cobranças de TODOS os clientes...")
        discrepancias_cobrancas = []
        clientes_com_divergencia = []
        
        # Validar TODOS os clientes do banco local
        clientes_para_validar = AsaasClienteSyncronizado.objects.all()
        total_clientes_para_validar = clientes_para_validar.count()
        
        logger.info(f"Validando cobranças de {total_clientes_para_validar} clientes...")
        
        for i, cliente in enumerate(clientes_para_validar, 1):
            if i % 50 == 0:  # Log a cada 50 clientes
                logger.info(f"Validado {i}/{total_clientes_para_validar} clientes...")
            
            # Buscar cobranças do ASAAS
            response_cobrancas = sync_service._fazer_requisicao(
                'GET', 
                'payments', 
                params={'customer': cliente.asaas_customer_id, 'limit': 100}
            )
            
            if response_cobrancas:
                total_cobrancas_asaas = response_cobrancas.get('totalCount', 0)
                cobrancas_asaas = response_cobrancas.get('data', [])
                
                # Contar cobranças no banco local
                total_cobrancas_local = cliente.cobrancas.count()
                
                if total_cobrancas_asaas != total_cobrancas_local:
                    # Identificar cobranças faltando
                    cobrancas_faltando = []
                    
                    for cobranca_data in cobrancas_asaas:
                        cobranca_id = cobranca_data.get('id')
                        if not AsaasCobrancaSyncronizada.objects.filter(asaas_payment_id=cobranca_id).exists():
                            cobrancas_faltando.append({
                                'asaas_id': cobranca_id,
                                'valor': cobranca_data.get('value', 0),
                                'status': cobranca_data.get('status', 'N/A'),
                                'vencimento': cobranca_data.get('dueDate', 'N/A'),
                                'descricao': cobranca_data.get('description', 'N/A'),
                            })
                    
                    clientes_com_divergencia.append({
                        'cliente': cliente,
                        'total_asaas': total_cobrancas_asaas,
                        'total_local': total_cobrancas_local,
                        'diferenca': total_cobrancas_asaas - total_cobrancas_local,
                        'cobrancas_faltando': cobrancas_faltando,
                    })
        
        logger.info(f"✅ Validação completa: {len(clientes_com_divergencia)} clientes com divergências em cobranças")
        
        # 3. ESTATÍSTICAS GERAIS
        total_cobrancas_local = AsaasCobrancaSyncronizada.objects.count()
        total_clientes_validados = total_clientes_para_validar
        
        # Calcular percentual de sincronização
        percentual_clientes = (total_clientes_local / total_clientes_asaas * 100) if total_clientes_asaas > 0 else 0
        
        context = {
            'usar_alternativo': usar_alternativo,
            'total_clientes_asaas': total_clientes_asaas,
            'total_clientes_local': total_clientes_local,
            'percentual_clientes': round(percentual_clientes, 2),
            'clientes_faltando': clientes_faltando,
            'total_clientes_faltando': len(clientes_faltando),
            'clientes_com_divergencia': clientes_com_divergencia,
            'total_cobrancas_local': total_cobrancas_local,
            'total_clientes_validados': total_clientes_validados,
            'data_validacao': timezone.now(),
        }
        
        return render(request, 'asaas_sync/validacao.html', context)
        
    except Exception as e:
        logger.error(f"Erro na validação: {str(e)}", exc_info=True)
        return render(request, 'asaas_sync/validacao.html', {
            'erro': f'Erro ao validar sincronização: {str(e)}'
        })


@login_required
def exportar_clientes_excel(request):
    """
    Exporta clientes para Excel.
    origem=local: Dados do banco de dados local
    origem=asaas: Dados direto da API ASAAS (padrão)
    """
    usar_alternativo = request.GET.get('alternativo', 'false') == 'true'
    origem = request.GET.get('origem', 'asaas')  # 'local' ou 'asaas'
    
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Clientes"
        
        # Cabeçalhos
        headers = [
            'ID ASAAS', 'Nome', 'CPF/CNPJ', 'Email', 'Telefone', 'Celular',
            'CEP', 'Endereço', 'Número', 'Complemento', 'Bairro', 'Cidade', 'Estado',
            'Inscrição Municipal', 'Inscrição Estadual', 'Observações',
            'Data Criação', 'Notificações Desabilitadas', 'Sincronizado Em'
        ]
        
        # Estilizar cabeçalho
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        row_num = 2
        total_exportados = 0
        
        if origem == 'local':
            # EXPORTAR DO BANCO DE DADOS LOCAL
            logger.info("Exportando clientes do banco de dados local...")
            clientes = AsaasClienteSyncronizado.objects.all().order_by('nome')
            
            for cliente in clientes:
                ws.cell(row=row_num, column=1, value=cliente.asaas_customer_id)
                ws.cell(row=row_num, column=2, value=cliente.nome)
                ws.cell(row=row_num, column=3, value=cliente.cpf_cnpj or '')
                ws.cell(row=row_num, column=4, value=cliente.email or '')
                ws.cell(row=row_num, column=5, value=cliente.telefone or '')
                ws.cell(row=row_num, column=6, value=cliente.celular or '')
                ws.cell(row=row_num, column=7, value=cliente.cep or '')
                ws.cell(row=row_num, column=8, value=cliente.endereco or '')
                ws.cell(row=row_num, column=9, value=cliente.numero or '')
                ws.cell(row=row_num, column=10, value=cliente.complemento or '')
                ws.cell(row=row_num, column=11, value=cliente.bairro or '')
                ws.cell(row=row_num, column=12, value=cliente.cidade or '')
                ws.cell(row=row_num, column=13, value=cliente.estado or '')
                ws.cell(row=row_num, column=14, value=cliente.inscricao_municipal or '')
                ws.cell(row=row_num, column=15, value=cliente.inscricao_estadual or '')
                ws.cell(row=row_num, column=16, value=cliente.observacoes or '')
                ws.cell(row=row_num, column=17, value=cliente.data_criacao_asaas.strftime('%d/%m/%Y %H:%M') if cliente.data_criacao_asaas else '')
                ws.cell(row=row_num, column=18, value='Sim' if cliente.notificacoes_desabilitadas else 'Não')
                ws.cell(row=row_num, column=19, value=cliente.sincronizado_em.strftime('%d/%m/%Y %H:%M') if cliente.sincronizado_em else '')
                
                row_num += 1
                total_exportados += 1
            
        else:
            # EXPORTAR DA API ASAAS
            sync_service = AsaasSyncService()
            if usar_alternativo:
                token_alternativo = getattr(settings, 'ASAAS_ALTERNATIVO_TOKEN', None)
                if token_alternativo:
                    sync_service.api_token = token_alternativo
                    sync_service.headers['access_token'] = token_alternativo
            
            logger.info("Exportando clientes da API ASAAS...")
            offset = 0
            limit = 100
            
            while True:
                response = sync_service._fazer_requisicao('GET', 'customers', params={'limit': limit, 'offset': offset})
                
                if not response or not response.get('data'):
                    break
                
                clientes = response.get('data', [])
                
                for cliente in clientes:
                    ws.cell(row=row_num, column=1, value=cliente.get('id', ''))
                    ws.cell(row=row_num, column=2, value=cliente.get('name', ''))
                    ws.cell(row=row_num, column=3, value=cliente.get('cpfCnpj', ''))
                    ws.cell(row=row_num, column=4, value=cliente.get('email', ''))
                    ws.cell(row=row_num, column=5, value=cliente.get('phone', ''))
                    ws.cell(row=row_num, column=6, value=cliente.get('mobilePhone', ''))
                    ws.cell(row=row_num, column=7, value=cliente.get('postalCode', ''))
                    ws.cell(row=row_num, column=8, value=cliente.get('address', ''))
                    ws.cell(row=row_num, column=9, value=cliente.get('addressNumber', ''))
                    ws.cell(row=row_num, column=10, value=cliente.get('complement', ''))
                    ws.cell(row=row_num, column=11, value=cliente.get('province', ''))
                    ws.cell(row=row_num, column=12, value=cliente.get('city', ''))
                    ws.cell(row=row_num, column=13, value=cliente.get('state', ''))
                    ws.cell(row=row_num, column=14, value=cliente.get('municipalInscription', ''))
                    ws.cell(row=row_num, column=15, value=cliente.get('stateInscription', ''))
                    ws.cell(row=row_num, column=16, value=cliente.get('observations', ''))
                    ws.cell(row=row_num, column=17, value=cliente.get('dateCreated', ''))
                    ws.cell(row=row_num, column=18, value='Sim' if cliente.get('notificationDisabled') else 'Não')
                    ws.cell(row=row_num, column=19, value=timezone.now().strftime('%d/%m/%Y %H:%M'))
                    
                    row_num += 1
                    total_exportados += 1
                
                if not response.get('hasMore', False):
                    break
                
                offset += limit
                
                if offset >= 10000:
                    logger.warning("Limite de 10000 clientes atingido")
                    break
        
        # Ajustar largura das colunas
        for col_num in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_num)].width = 20
        
        # Preparar response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        origem_nome = 'local' if origem == 'local' else ('asaas_alternativo' if usar_alternativo else 'asaas_principal')
        nome_arquivo = f"clientes_{origem_nome}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{nome_arquivo}"'
        
        wb.save(response)
        
        logger.info(f"✅ Exportados {total_exportados} clientes ({origem}) para Excel")
        return response
        
    except Exception as e:
        logger.error(f"Erro ao exportar clientes: {str(e)}", exc_info=True)
        return HttpResponse(f"Erro ao exportar: {str(e)}", status=500)


@login_required
def exportar_cobrancas_excel(request):
    """
    Exporta cobranças para Excel.
    origem=local: Dados do banco de dados local
    origem=asaas: Dados direto da API ASAAS (padrão)
    """
    usar_alternativo = request.GET.get('alternativo', 'false') == 'true'
    origem = request.GET.get('origem', 'asaas')
    cliente_id = request.GET.get('cliente_id', None)
    
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Cobranças"
        
        # Cabeçalhos
        headers = [
            'ID Cobrança', 'ID Cliente', 'Nome Cliente', 'Valor', 'Valor Líquido',
            'Status', 'Descrição', 'Data Vencimento', 'Data Pagamento',
            'Forma Pagamento', 'Referência Externa', 'Invoice URL',
            'Bank Slip URL', 'Data Criação', 'PIX Copia e Cola'
        ]
        
        # Estilizar cabeçalho
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        row_num = 2
        total_exportados = 0
        
        if origem == 'local':
            # EXPORTAR DO BANCO DE DADOS LOCAL
            logger.info("Exportando cobranças do banco de dados local...")
            cobrancas = AsaasCobrancaSyncronizada.objects.all().order_by('-data_vencimento')
            
            if cliente_id:
                cobrancas = cobrancas.filter(cliente__asaas_customer_id=cliente_id)
            
            for cobranca in cobrancas:
                ws.cell(row=row_num, column=1, value=cobranca.asaas_payment_id)
                ws.cell(row=row_num, column=2, value=cobranca.cliente.asaas_customer_id if cobranca.cliente else '')
                ws.cell(row=row_num, column=3, value=cobranca.cliente.nome if cobranca.cliente else '')
                ws.cell(row=row_num, column=4, value=float(cobranca.valor))
                ws.cell(row=row_num, column=5, value=float(cobranca.valor_liquido) if cobranca.valor_liquido else 0)
                ws.cell(row=row_num, column=6, value=cobranca.status)
                ws.cell(row=row_num, column=7, value=cobranca.descricao or '')
                ws.cell(row=row_num, column=8, value=cobranca.data_vencimento.strftime('%d/%m/%Y') if cobranca.data_vencimento else '')
                ws.cell(row=row_num, column=9, value=cobranca.data_pagamento.strftime('%d/%m/%Y') if cobranca.data_pagamento else '')
                ws.cell(row=row_num, column=10, value=cobranca.tipo_cobranca or '')
                ws.cell(row=row_num, column=11, value=cobranca.external_reference or '')
                ws.cell(row=row_num, column=12, value=cobranca.invoice_url or '')
                ws.cell(row=row_num, column=13, value=cobranca.bank_slip_url or '')
                ws.cell(row=row_num, column=14, value=cobranca.data_criacao_asaas.strftime('%d/%m/%Y') if cobranca.data_criacao_asaas else '')
                ws.cell(row=row_num, column=15, value=cobranca.pix_copy_paste or '')
                
                row_num += 1
                total_exportados += 1
                
        else:
            # EXPORTAR DA API ASAAS
            sync_service = AsaasSyncService()
            if usar_alternativo:
                token_alternativo = getattr(settings, 'ASAAS_ALTERNATIVO_TOKEN', None)
                if token_alternativo:
                    sync_service.api_token = token_alternativo
                    sync_service.headers['access_token'] = token_alternativo
            
            logger.info("Exportando cobranças da API ASAAS...")
            offset = 0
            limit = 100
            
            params = {'limit': limit, 'offset': offset}
            if cliente_id:
                params['customer'] = cliente_id
            
            while True:
                response = sync_service._fazer_requisicao('GET', 'payments', params=params)
                
                if not response or not response.get('data'):
                    break
                
                cobrancas = response.get('data', [])
                
                for cobranca in cobrancas:
                    ws.cell(row=row_num, column=1, value=cobranca.get('id', ''))
                    ws.cell(row=row_num, column=2, value=cobranca.get('customer', ''))
                    ws.cell(row=row_num, column=3, value=cobranca.get('customerName', ''))
                    ws.cell(row=row_num, column=4, value=cobranca.get('value', 0))
                    ws.cell(row=row_num, column=5, value=cobranca.get('netValue', 0))
                    ws.cell(row=row_num, column=6, value=cobranca.get('status', ''))
                    ws.cell(row=row_num, column=7, value=cobranca.get('description', ''))
                    ws.cell(row=row_num, column=8, value=cobranca.get('dueDate', ''))
                    ws.cell(row=row_num, column=9, value=cobranca.get('paymentDate', ''))
                    ws.cell(row=row_num, column=10, value=cobranca.get('billingType', ''))
                    ws.cell(row=row_num, column=11, value=cobranca.get('nossoNumero', ''))
                    ws.cell(row=row_num, column=12, value=cobranca.get('invoiceUrl', ''))
                    ws.cell(row=row_num, column=13, value=cobranca.get('bankSlipUrl', ''))
                    ws.cell(row=row_num, column=14, value=cobranca.get('dateCreated', ''))
                    ws.cell(row=row_num, column=15, value='Sim' if cobranca.get('confirmed') else 'Não')
                    
                    row_num += 1
                    total_exportados += 1
                
                if not response.get('hasMore', False):
                    break
                
                offset += limit
                params['offset'] = offset
                
                if offset >= 10000:
                    logger.warning("Limite de 10000 cobranças atingido")
                    break
        
        # Ajustar largura das colunas
        for col_num in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_num)].width = 20
        
        # Preparar response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        origem_nome = 'local' if origem == 'local' else ('asaas_alternativo' if usar_alternativo else 'asaas_principal')
        nome_arquivo = f"cobrancas_{origem_nome}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{nome_arquivo}"'
        
        wb.save(response)
        
        logger.info(f"✅ Exportadas {total_exportados} cobranças ({origem}) para Excel")
        return response
        
    except Exception as e:
        logger.error(f"Erro ao exportar cobranças: {str(e)}", exc_info=True)
        return HttpResponse(f"Erro ao exportar: {str(e)}", status=500)


@login_required
def exportar_clientes_com_boletos_excel(request):
    """
    Exporta relatório completo: Clientes com todos os seus boletos.
    Cada cliente tem suas cobranças listadas abaixo dele.
    origem=local: Dados do banco de dados local
    origem=asaas: Dados direto da API ASAAS (padrão)
    """
    usar_alternativo = request.GET.get('alternativo', 'false') == 'true'
    origem = request.GET.get('origem', 'asaas')
    
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Clientes e Boletos"
        
        # Configurar service se for API
        sync_service = None
        if origem == 'asaas':
            sync_service = AsaasSyncService()
            if usar_alternativo:
                token_alternativo = getattr(settings, 'ASAAS_ALTERNATIVO_TOKEN', None)
                if token_alternativo:
                    sync_service.api_token = token_alternativo
                    sync_service.headers['access_token'] = token_alternativo
        
        row_num = 1
        total_clientes = 0
        total_boletos = 0
        
        if origem == 'local':
            # EXPORTAR DO BANCO DE DADOS LOCAL
            logger.info("Exportando clientes com boletos do banco local...")
            clientes = AsaasClienteSyncronizado.objects.all().order_by('nome')
            
            for cliente in clientes:
                # CABEÇALHO DO CLIENTE
                ws.cell(row=row_num, column=1, value="CLIENTE")
                ws.cell(row=row_num, column=1).font = Font(bold=True, size=12, color="FFFFFF")
                ws.cell(row=row_num, column=1).fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                ws.merge_cells(f'A{row_num}:G{row_num}')
                row_num += 1
                
                # Dados do cliente
                ws.cell(row=row_num, column=1, value="ID ASAAS:")
                ws.cell(row=row_num, column=2, value=cliente.asaas_customer_id)
                row_num += 1
                
                ws.cell(row=row_num, column=1, value="Nome:")
                ws.cell(row=row_num, column=2, value=cliente.nome)
                ws.cell(row=row_num, column=3, value="CPF/CNPJ:")
                ws.cell(row=row_num, column=4, value=cliente.cpf_cnpj or '')
                row_num += 1
                
                ws.cell(row=row_num, column=1, value="Email:")
                ws.cell(row=row_num, column=2, value=cliente.email or '')
                ws.cell(row=row_num, column=3, value="Telefone:")
                ws.cell(row=row_num, column=4, value=cliente.telefone or '')
                row_num += 1
                
                ws.cell(row=row_num, column=1, value="Endereço:")
                endereco_completo = f"{cliente.endereco or ''}, {cliente.numero or ''} - {cliente.bairro or ''}, {cliente.cidade or ''}/{cliente.estado or ''}"
                ws.cell(row=row_num, column=2, value=endereco_completo)
                row_num += 1
                
                # COBRANÇAS DO CLIENTE
                cobrancas = AsaasCobrancaSyncronizada.objects.filter(cliente=cliente).order_by('-data_vencimento')
                
                if cobrancas.exists():
                    row_num += 1
                    ws.cell(row=row_num, column=1, value="BOLETOS/COBRANÇAS")
                    ws.cell(row=row_num, column=1).font = Font(bold=True, color="FFFFFF")
                    ws.cell(row=row_num, column=1).fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                    ws.merge_cells(f'A{row_num}:G{row_num}')
                    row_num += 1
                    
                    # Cabeçalhos das cobranças
                    headers_cobranca = ['ID', 'Valor', 'Status', 'Vencimento', 'Pagamento', 'Descrição', 'Boleto URL']
                    for col_num, header in enumerate(headers_cobranca, 1):
                        cell = ws.cell(row=row_num, column=col_num)
                        cell.value = header
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                    row_num += 1
                    
                    # Dados das cobranças
                    for cobranca in cobrancas:
                        ws.cell(row=row_num, column=1, value=cobranca.asaas_payment_id)
                        ws.cell(row=row_num, column=2, value=f"R$ {float(cobranca.valor):.2f}")
                        ws.cell(row=row_num, column=3, value=cobranca.status)
                        ws.cell(row=row_num, column=4, value=cobranca.data_vencimento.strftime('%d/%m/%Y') if cobranca.data_vencimento else '')
                        ws.cell(row=row_num, column=5, value=cobranca.data_pagamento.strftime('%d/%m/%Y') if cobranca.data_pagamento else '')
                        ws.cell(row=row_num, column=6, value=cobranca.descricao or '')
                        ws.cell(row=row_num, column=7, value=cobranca.bank_slip_url or '')
                        row_num += 1
                        total_boletos += 1
                else:
                    row_num += 1
                    ws.cell(row=row_num, column=1, value="Nenhuma cobrança encontrada")
                    ws.cell(row=row_num, column=1).font = Font(italic=True, color="999999")
                    row_num += 1
                
                # Linha em branco separadora
                row_num += 1
                total_clientes += 1
                
        else:
            # EXPORTAR DA API ASAAS
            logger.info("Exportando clientes com boletos da API ASAAS...")
            offset = 0
            limit = 50  # Menos clientes porque faremos muitas requisições
            
            while True:
                response_clientes = sync_service._fazer_requisicao('GET', 'customers', params={'limit': limit, 'offset': offset})
                
                if not response_clientes or not response_clientes.get('data'):
                    break
                
                clientes = response_clientes.get('data', [])
                
                for cliente in clientes:
                    # CABEÇALHO DO CLIENTE
                    ws.cell(row=row_num, column=1, value="CLIENTE")
                    ws.cell(row=row_num, column=1).font = Font(bold=True, size=12, color="FFFFFF")
                    ws.cell(row=row_num, column=1).fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    ws.merge_cells(f'A{row_num}:G{row_num}')
                    row_num += 1
                    
                    # Dados do cliente
                    ws.cell(row=row_num, column=1, value="ID ASAAS:")
                    ws.cell(row=row_num, column=2, value=cliente.get('id', ''))
                    row_num += 1
                    
                    ws.cell(row=row_num, column=1, value="Nome:")
                    ws.cell(row=row_num, column=2, value=cliente.get('name', ''))
                    ws.cell(row=row_num, column=3, value="CPF/CNPJ:")
                    ws.cell(row=row_num, column=4, value=cliente.get('cpfCnpj', ''))
                    row_num += 1
                    
                    ws.cell(row=row_num, column=1, value="Email:")
                    ws.cell(row=row_num, column=2, value=cliente.get('email', ''))
                    ws.cell(row=row_num, column=3, value="Telefone:")
                    ws.cell(row=row_num, column=4, value=cliente.get('mobilePhone', '') or cliente.get('phone', ''))
                    row_num += 1
                    
                    ws.cell(row=row_num, column=1, value="Endereço:")
                    endereco_completo = f"{cliente.get('address', '')}, {cliente.get('addressNumber', '')} - {cliente.get('province', '')}, {cliente.get('city', '')}/{cliente.get('state', '')}"
                    ws.cell(row=row_num, column=2, value=endereco_completo)
                    row_num += 1
                    
                    # Buscar cobranças deste cliente
                    cliente_id = cliente.get('id')
                    response_cobrancas = sync_service._fazer_requisicao('GET', 'payments', params={'customer': cliente_id, 'limit': 100})
                    
                    if response_cobrancas and response_cobrancas.get('data'):
                        cobrancas = response_cobrancas.get('data', [])
                        
                        row_num += 1
                        ws.cell(row=row_num, column=1, value="BOLETOS/COBRANÇAS")
                        ws.cell(row=row_num, column=1).font = Font(bold=True, color="FFFFFF")
                        ws.cell(row=row_num, column=1).fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                        ws.merge_cells(f'A{row_num}:G{row_num}')
                        row_num += 1
                        
                        # Cabeçalhos das cobranças
                        headers_cobranca = ['ID', 'Valor', 'Status', 'Vencimento', 'Pagamento', 'Descrição', 'Boleto URL']
                        for col_num, header in enumerate(headers_cobranca, 1):
                            cell = ws.cell(row=row_num, column=col_num)
                            cell.value = header
                            cell.font = Font(bold=True)
                            cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                        row_num += 1
                        
                        # Dados das cobranças
                        for cobranca in cobrancas:
                            ws.cell(row=row_num, column=1, value=cobranca.get('id', ''))
                            ws.cell(row=row_num, column=2, value=f"R$ {cobranca.get('value', 0):.2f}")
                            ws.cell(row=row_num, column=3, value=cobranca.get('status', ''))
                            ws.cell(row=row_num, column=4, value=cobranca.get('dueDate', ''))
                            ws.cell(row=row_num, column=5, value=cobranca.get('paymentDate', ''))
                            ws.cell(row=row_num, column=6, value=cobranca.get('description', ''))
                            ws.cell(row=row_num, column=7, value=cobranca.get('bankSlipUrl', ''))
                            row_num += 1
                            total_boletos += 1
                    else:
                        row_num += 1
                        ws.cell(row=row_num, column=1, value="Nenhuma cobrança encontrada")
                        ws.cell(row=row_num, column=1).font = Font(italic=True, color="999999")
                        row_num += 1
                    
                    # Linha em branco separadora
                    row_num += 1
                    total_clientes += 1
                
                if not response_clientes.get('hasMore', False):
                    break
                
                offset += limit
                
                # Limite de segurança (50 clientes = muitas requisições)
                if offset >= 500:
                    logger.warning("Limite de 500 clientes atingido na exportação completa")
                    break
        
        # Ajustar largura das colunas
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 40
        ws.column_dimensions['G'].width = 50
        
        # Adicionar resumo no topo
        ws.insert_rows(1, 3)
        ws.cell(row=1, column=1, value="RELATÓRIO COMPLETO - CLIENTES E BOLETOS")
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)
        ws.merge_cells('A1:G1')
        
        ws.cell(row=2, column=1, value=f"Total de Clientes: {total_clientes}")
        ws.cell(row=2, column=3, value=f"Total de Boletos: {total_boletos}")
        ws.cell(row=3, column=1, value=f"Data da Exportação: {timezone.now().strftime('%d/%m/%Y %H:%M')}")
        
        # Preparar response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        origem_nome = 'local' if origem == 'local' else ('asaas_alternativo' if usar_alternativo else 'asaas_principal')
        nome_arquivo = f"clientes_boletos_{origem_nome}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{nome_arquivo}"'
        
        wb.save(response)
        
        logger.info(f"✅ Exportados {total_clientes} clientes com {total_boletos} boletos ({origem})")
        return response
        
    except Exception as e:
        logger.error(f"Erro ao exportar relatório completo: {str(e)}", exc_info=True)
        return HttpResponse(f"Erro ao exportar: {str(e)}", status=500)
        
        # Cabeçalhos
        headers = [
            'ID Cobrança', 'ID Cliente', 'Nome Cliente', 'Valor', 'Valor Líquido',
            'Status', 'Descrição', 'Data Vencimento', 'Data Pagamento',
            'Forma Pagamento', 'Nosso Número', 'Invoice URL',
            'Bank Slip URL', 'Data Criação', 'Confirmado'
        ]
        
        # Estilizar cabeçalho
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Buscar cobranças
        logger.info("Buscando cobranças do ASAAS para exportação...")
        offset = 0
        limit = 100
        row_num = 2
        total_exportados = 0
        
        params = {'limit': limit, 'offset': offset}
        if cliente_id:
            params['customer'] = cliente_id
        
        while True:
            response = sync_service._fazer_requisicao('GET', 'payments', params=params)
            
            if not response or not response.get('data'):
                break
            
            cobrancas = response.get('data', [])
            
            for cobranca in cobrancas:
                ws.cell(row=row_num, column=1, value=cobranca.get('id', ''))
                ws.cell(row=row_num, column=2, value=cobranca.get('customer', ''))
                ws.cell(row=row_num, column=3, value=cobranca.get('customerName', ''))
                ws.cell(row=row_num, column=4, value=cobranca.get('value', 0))
                ws.cell(row=row_num, column=5, value=cobranca.get('netValue', 0))
                ws.cell(row=row_num, column=6, value=cobranca.get('status', ''))
                ws.cell(row=row_num, column=7, value=cobranca.get('description', ''))
                ws.cell(row=row_num, column=8, value=cobranca.get('dueDate', ''))
                ws.cell(row=row_num, column=9, value=cobranca.get('paymentDate', ''))
                ws.cell(row=row_num, column=10, value=cobranca.get('billingType', ''))
                ws.cell(row=row_num, column=11, value=cobranca.get('nossoNumero', ''))
                ws.cell(row=row_num, column=12, value=cobranca.get('invoiceUrl', ''))
                ws.cell(row=row_num, column=13, value=cobranca.get('bankSlipUrl', ''))
                ws.cell(row=row_num, column=14, value=cobranca.get('dateCreated', ''))
                ws.cell(row=row_num, column=15, value='Sim' if cobranca.get('confirmed') else 'Não')
                
                row_num += 1
                total_exportados += 1
            
            # Verificar se há mais páginas
            if not response.get('hasMore', False):
                break
            
            offset += limit
            params['offset'] = offset
            
            # Limite de segurança
            if offset >= 10000:
                logger.warning("Limite de 10000 cobranças atingido na exportação")
                break
        
        # Ajustar largura das colunas
        for col_num in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_num)
            ws.column_dimensions[column_letter].width = 20
        
        # Preparar response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        nome_arquivo = f"cobrancas_asaas_{'alternativo' if usar_alternativo else 'principal'}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{nome_arquivo}"'
        
        wb.save(response)
        
        logger.info(f"✅ Exportadas {total_exportados} cobranças para Excel")
        return response
        
    except Exception as e:
        logger.error(f"Erro ao exportar cobranças: {str(e)}", exc_info=True)
        return HttpResponse(f"Erro ao exportar: {str(e)}", status=500)